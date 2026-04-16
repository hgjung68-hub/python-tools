import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import pyodbc
import pandas as pd
from datetime import datetime, date
import os
from openpyxl.styles import Alignment, PatternFill, Font
 
 
# ─────────────────────────────────────────────
#  설정 구간
# ─────────────────────────────────────────────
DEFAULT_MDB_PATH = r"C:\Program Files (x86)\KTT_FPReader\dbFolder\hds_fpsystem.mdb"

# 시도할 암호 목록 (분석하신 후보 암호들을 순서대로 넣습니다)
MDB_PWD_CANDIDATES = [
    r"SysTools@365#CL",
    r"tjdgustltmxpa", 
    r""                 # 암호 없음 대비
]
 
EMP_NUM_MAP = {
    "3050019": "1230084",
    # "3050020": "1230085",  # 추가 변환이 필요한 경우 계속 추가
}
# ─────────────────────────────────────────────
 
def try_connect(mdb_path: str, log) -> "pyodbc.Connection":
    """
    여러 개의 암호 후보를 순차적으로 시도하여 연결을 시도
    """
    last_exc = None
    
    for pwd in MDB_PWD_CANDIDATES:
        display_pwd = "*" * len(pwd) if pwd else "None"
        log(f"연결 시도 중... (PWD: {display_pwd})")
        
        # ODBC 연결 문자열 구성
        conn_str = (
            f"Driver={{Microsoft Access Driver (*.mdb, *.accdb)}};"
            f"DBQ={mdb_path};"
        )
        if pwd:
            conn_str += f"PWD={pwd};"

        try:
            conn = pyodbc.connect(conn_str, timeout=5)
            log(f"  → 연결 성공!")
            return conn
        except Exception as exc:
            log(f"  → 인증 실패")
            last_exc = exc

    # 모든 후보 실패 시
    raise RuntimeError(
        f"MDB 모든 암호 인증 실패.\n"
        f"시도된 암호 개수: {len(MDB_PWD_CANDIDATES)}개\n"
        f"최종 오류: {last_exc}"
    )
 
def generate_excel(mdb_path, start_date, end_date, log, done_callback):
    """백그라운드 스레드에서 실행되는 메인 처리 함수"""
    try:
        log("MDB 연결 중...")
        conn = try_connect(mdb_path, log)
        query = f"""
            SELECT
                date_Attestation,
                str_Week,
                str_workempNum,
                str_workempName,
                str_Mode
            FROM tb_workresult
            WHERE date_Attestation >= #{start_date} 00:00:00#
              AND date_Attestation <= #{end_date} 23:59:59#
              AND Len(str_workempNum) = 7
        """
        log("데이터 조회 중...")
        df = pd.read_sql(query, conn)
        conn.close()
 
        if df.empty:
            log(f"[!] {start_date} ~ {end_date} 기간에 조건에 맞는 데이터가 없습니다.")
            done_callback(success=False)
            return
 
        df['date_Attestation'] = pd.to_datetime(df['date_Attestation'])
        df['only_date'] = df['date_Attestation'].dt.date
 
        first_in  = df.loc[df.groupby(['str_workempNum', 'only_date'])['date_Attestation'].idxmin()]
        last_out  = df.loc[df.groupby(['str_workempNum', 'only_date'])['date_Attestation'].idxmax()]
 
        final_df = (
            pd.concat([first_in, last_out])
            .drop_duplicates()
            .sort_values(by='date_Attestation')
        )
 
        export_df = final_df[[
            'date_Attestation', 'str_Week', 'str_workempNum', 'str_workempName', 'str_Mode'
        ]].copy()
 
        export_df['date_Attestation'] = export_df['date_Attestation'].dt.strftime('%Y-%m-%d %H:%M:%S')
        export_df.columns = ['인증일시', '요일', '사원번호', '이름', '인증모드']
 
        changed = export_df['사원번호'].map(EMP_NUM_MAP)
        converted_count = changed.notna().sum()
        export_df['사원번호'] = export_df['사원번호'].map(EMP_NUM_MAP).fillna(export_df['사원번호'])
 
        output_file = f"지문_출퇴근기록_{start_date}_{end_date}.xlsx"
        log(f"엑셀 저장 중: {output_file}")
 
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            export_df.to_excel(writer, index=False, sheet_name='Sheet1')
            ws = writer.sheets['Sheet1']
 
            center = Alignment(horizontal='center', vertical='center')
            h_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            h_font = Font(bold=True)
 
            for col, width in {'A': 22, 'B': 6, 'C': 10, 'D': 15, 'E': 10}.items():
                ws.column_dimensions[col].width = width
 
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = center
                    if cell.row == 1:
                        cell.fill = h_fill
                        cell.font = h_font
 
        total   = len(export_df)
        emp_cnt = export_df['사원번호'].nunique()
 
        log("─" * 48)
        log(f"[완료] 파일 생성: {os.path.abspath(output_file)}")
        log(f"  ▶ Data 생성 직원 : {emp_cnt:,} 명")
        log(f"  ▶ 총 데이터 건수 : {total:,} 건")
        if converted_count > 0:
            log(f"  ▶ 사원번호 변환  : {converted_count:,} 건 (매핑 테이블 적용)")
        log("─" * 48)
 
        done_callback(success=True, filepath=os.path.abspath(output_file))
 
    except Exception as e:
        log(f"\n[오류] {e}")
        done_callback(success=False)
 
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("TiSS용 KT텔레캅 지문인식기 출퇴근 데이터 추출")
        self.resizable(False, False)
        self._build_ui()
        self._center_window()
 
    # ── UI 구성 ──────────────────────────────────
    def _build_ui(self):
        PAD = dict(padx=10, pady=6)
 
        # ── MDB 파일 경로 ──────────────────────
        frame_mdb = ttk.LabelFrame(self, text=" MDB 파일 설정 ", padding=8)
        frame_mdb.grid(row=0, column=0, columnspan=3, sticky="ew", padx=14, pady=(14, 4))
 
        ttk.Label(frame_mdb, text="MDB 경로:").grid(row=0, column=0, sticky="w")
        self.var_path = tk.StringVar(value=DEFAULT_MDB_PATH)
        ttk.Entry(frame_mdb, textvariable=self.var_path, width=54).grid(
            row=0, column=1, padx=6)
        ttk.Button(frame_mdb, text="찾아보기", command=self._browse_mdb).grid(
            row=0, column=2)
 
        # ── 날짜 선택 ──────────────────────────
        frame_date = ttk.LabelFrame(self, text=" 추출 기간 ", padding=8)
        frame_date.grid(row=1, column=0, columnspan=3, sticky="ew", padx=14, pady=4)
 
        today = date.today()
 
        ttk.Label(frame_date, text="시작 일자:").grid(row=0, column=0, sticky="w")
        self.var_start = tk.StringVar(value=today.replace(day=1).strftime("%Y-%m-%d"))
        ttk.Entry(frame_date, textvariable=self.var_start, width=14).grid(
            row=0, column=1, padx=6)
        ttk.Label(frame_date, text="(예: 2026-04-01)", foreground="gray").grid(
            row=0, column=2, sticky="w")
 
        ttk.Label(frame_date, text="종료 일자:").grid(row=1, column=0, sticky="w", pady=(6, 0))
        self.var_end = tk.StringVar(value=today.strftime("%Y-%m-%d"))
        ttk.Entry(frame_date, textvariable=self.var_end, width=14).grid(
            row=1, column=1, padx=6, pady=(6, 0))
        ttk.Label(frame_date, text="(예: 2026-04-30)", foreground="gray").grid(
            row=1, column=2, sticky="w", pady=(6, 0))
 
        # ── 실행 버튼 & 진행 표시 ─────────────
        frame_btn = ttk.Frame(self)
        frame_btn.grid(row=2, column=0, columnspan=3, padx=14, pady=8, sticky="ew")
 
        self.btn_run = ttk.Button(
            frame_btn, text="▶  추출 실행", command=self._run, width=20)
        self.btn_run.pack(side="left")
 
        self.progress = ttk.Progressbar(frame_btn, mode="indeterminate", length=200)
        self.progress.pack(side="left", padx=(14, 0))
 
        # ── 로그 창 ───────────────────────────
        frame_log = ttk.LabelFrame(self, text=" 처리 로그 ", padding=6)
        frame_log.grid(row=3, column=0, columnspan=3, sticky="nsew",
                       padx=14, pady=(0, 14))
 
        self.log_text = tk.Text(
            frame_log, height=12, width=70,
            state="disabled", font=("Consolas", 9))
        scroll = ttk.Scrollbar(frame_log, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scroll.set)
        self.log_text.grid(row=0, column=0, sticky="nsew")
        scroll.grid(row=0, column=1, sticky="ns")
 
    def _center_window(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        x = (self.winfo_screenwidth()  - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"+{x}+{y}")
 
    # ── 이벤트 핸들러 ─────────────────────────
    def _browse_mdb(self):
        path = filedialog.askopenfilename(
            title="MDB 파일 선택",
            filetypes=[("Access Database", "*.mdb *.accdb"), ("All files", "*.*")]
        )
        if path:
            self.var_path.set(path)
 
    def _log(self, msg: str):
        """스레드-안전 로그 추가"""
        def _append():
            self.log_text.configure(state="normal")
            self.log_text.insert("end", msg + "\n")
            self.log_text.see("end")
            self.log_text.configure(state="disabled")
        self.after(0, _append)
 
    def _run(self):
        mdb_path   = self.var_path.get().strip()
        start_date = self.var_start.get().strip()
        end_date   = self.var_end.get().strip()
 
        # 간단한 유효성 검사
        if not mdb_path:
            messagebox.showerror("입력 오류", "MDB 경로를 입력하세요.")
            return
        try:
            datetime.strptime(start_date, "%Y-%m-%d")
            datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("입력 오류", "날짜 형식을 YYYY-MM-DD 로 입력하세요.")
            return
        if start_date > end_date:
            messagebox.showerror("입력 오류", "시작 일자가 종료 일자보다 늦습니다.")
            return
 
        self.btn_run.configure(state="disabled")
        self.progress.start(10)
        self._log(f"▶ 추출 시작: {start_date} ~ {end_date}")
 
        def done(success, filepath=None):
            self.after(0, self.progress.stop)
            self.after(0, lambda: self.btn_run.configure(state="normal"))
            if success:
                ans = messagebox.askyesno(
                    "완료",
                    f"파일이 생성되었습니다.\n{filepath}\n\n파일이 있는 폴더를 여시겠습니까?"
                )
                if ans:
                    os.startfile(os.path.dirname(filepath))
 
        threading.Thread(
            target=generate_excel,
            args=(mdb_path, start_date, end_date, self._log, done),
            daemon=True
        ).start()
  
if __name__ == "__main__":
    app = App()
    app.mainloop()